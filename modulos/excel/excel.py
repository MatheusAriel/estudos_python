"""
https://openpyxl.readthedocs.io/en/stable/
"""

import openpyxl
from random import uniform

# COMO LER A ALTERAR
"""
pedidos = openpyxl.load_workbook('pedidos.xlsx')
nome_planilhas = pedidos.sheetnames
#print(nome_planilhas)

planilha1 = pedidos['Página1']
#print(planilha1['B4'].value)
#print(planilha1['A'])
planilha1['B4'].value = 2400

# for campo in planilha1['a']:
#     print(campo.value)

# for linha in planilha1['a1:c4']:
#     for coluna in linha:
#         print(coluna.value)

# for linha in planilha1:
#     for coluna in linha:
#         print(coluna.value)

# for linha in planilha1:
#     print(linha[0].value, linha[1].value, linha[2].value)

for linha in range(5, 16):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 1500), 2)
    planilha1.cell(linha, 3).value = preco

pedidos.save('nova_planilha_modificada.xlsx')

"""

# COMO CRIAR

planilha = openpyxl.Workbook()
planilha.create_sheet('Página1', 0)
planilha.create_sheet('Página2', 1)

planilha1 = planilha['Página1']
planilha2 = planilha['Página2']

for linha in range(1, 11):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

for linha in range(1, 11):
    planilha2.cell(linha, 1).value = f'Luiz {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, 2).value = f'Otávio {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, 3).value = f'Joãozinho {linha} {round(uniform(10, 100), 2)}'

planilha.save('nova_planilha_criada.xlsx')
